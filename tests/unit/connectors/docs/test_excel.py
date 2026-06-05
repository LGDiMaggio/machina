"""Tests for ExcelCsvConnector — read, write, coercion edge cases."""

from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from machina.connectors.docs.excel import (
    COERCER_REGISTRY,
    ExcelCsvConnector,
    _bool_it,
    _coerce_cell,
    _date_parse,
    _datetime_parse,
    _excel_serial_to_date,
    _float_it,
    _guard_formula,
    _int_it,
    _strip_formula_guard,
)
from machina.connectors.docs.excel_schema import (
    ColumnMapping,
    ExcelConnectorConfig,
    SheetSchema,
)
from machina.domain.asset import AssetType, Criticality
from machina.domain.work_order import Priority, WorkOrder, WorkOrderStatus, WorkOrderType
from machina.exceptions import (
    ConnectorConfigError,
    ConnectorSchemaError,
)

FIXTURES = Path(__file__).resolve().parents[3] / "fixtures" / "excel"


# ------------------------------------------------------------------
# Coercer unit tests
# ------------------------------------------------------------------


class TestFloatIt:
    def test_int_input(self) -> None:
        assert _float_it(42) == 42.0

    def test_float_input(self) -> None:
        assert _float_it(3.14) == 3.14

    def test_italian_decimal_comma(self) -> None:
        assert _float_it("3,14") == 3.14

    def test_normal_dot(self) -> None:
        assert _float_it("3.14") == 3.14

    def test_whitespace(self) -> None:
        assert _float_it("  42.5  ") == 42.5

    def test_invalid_raises(self) -> None:
        with pytest.raises(ValueError):
            _float_it("not a number")


class TestIntIt:
    def test_int(self) -> None:
        assert _int_it(42) == 42

    def test_float_truncates(self) -> None:
        assert _int_it(3.9) == 3

    def test_string(self) -> None:
        assert _int_it("7") == 7


class TestDateParse:
    def test_italian_dd_mm_yyyy(self) -> None:
        assert _date_parse("15/03/2019") == date(2019, 3, 15)

    def test_iso_yyyy_mm_dd(self) -> None:
        assert _date_parse("2020-06-22") == date(2020, 6, 22)

    def test_dd_dot_mm_dot_yyyy(self) -> None:
        assert _date_parse("15.03.2019") == date(2019, 3, 15)

    def test_dd_dash_mm_dash_yyyy(self) -> None:
        assert _date_parse("15-03-2019") == date(2019, 3, 15)

    def test_excel_serial(self) -> None:
        assert _date_parse(43831) == date(2020, 1, 1)

    def test_datetime_input(self) -> None:
        dt = datetime(2020, 6, 22, 14, 30)
        assert _date_parse(dt) == date(2020, 6, 22)

    def test_date_input(self) -> None:
        d = date(2020, 6, 22)
        assert _date_parse(d) == d

    def test_invalid_raises(self) -> None:
        with pytest.raises((ValueError, TypeError)):
            _date_parse("not-a-date")


class TestDatetimeParse:
    def test_iso_string(self) -> None:
        result = _datetime_parse("2020-06-22T14:30:00")
        assert result.year == 2020
        assert result.month == 6
        assert result.tzinfo is not None

    def test_date_only_string(self) -> None:
        result = _datetime_parse("15/03/2019")
        assert result.date() == date(2019, 3, 15)

    def test_datetime_input(self) -> None:
        dt = datetime(2020, 1, 1, 12, 0)
        result = _datetime_parse(dt)
        assert result.tzinfo is not None


class TestExcelSerialToDate:
    def test_epoch(self) -> None:
        assert _excel_serial_to_date(1) == date(1899, 12, 31)

    def test_known_date(self) -> None:
        assert _excel_serial_to_date(43831) == date(2020, 1, 1)

    def test_recent_date(self) -> None:
        assert _excel_serial_to_date(45000) == date(2023, 3, 15)


class TestBoolIt:
    @pytest.mark.parametrize("val", ["1", "true", "yes", "sì", "si", "vero", "x", "True", "YES"])
    def test_truthy(self, val: str) -> None:
        assert _bool_it(val) is True

    @pytest.mark.parametrize("val", ["0", "false", "no", "falso", "", "False", "NO"])
    def test_falsy(self, val: str) -> None:
        assert _bool_it(val) is False

    def test_bool_passthrough(self) -> None:
        assert _bool_it(True) is True
        assert _bool_it(False) is False

    def test_invalid_raises(self) -> None:
        with pytest.raises(ValueError, match="Cannot coerce"):
            _bool_it("maybe")


class TestCoerceCell:
    def test_empty_string_non_required_returns_default(self) -> None:
        m = ColumnMapping(column="X", field="x", default="fallback")
        assert _coerce_cell("", m) == "fallback"

    def test_none_non_required_returns_default(self) -> None:
        m = ColumnMapping(column="X", field="x", default=0)
        assert _coerce_cell(None, m) == 0

    def test_empty_required_returns_none(self) -> None:
        m = ColumnMapping(column="X", field="x", required=True)
        assert _coerce_cell("", m) is None

    def test_named_coercer(self) -> None:
        m = ColumnMapping(column="X", field="x", coerce="float_it")
        assert _coerce_cell("3,14", m) == 3.14

    def test_type_coercion_fallback(self) -> None:
        m = ColumnMapping(column="X", field="x", type="float")
        assert _coerce_cell("42.5", m) == 42.5


class TestCoercerRegistry:
    def test_all_registered(self) -> None:
        expected = {
            "float_it",
            "int_it",
            "date_parse",
            "italian_date",
            "datetime_parse",
            "bool_it",
            "strip",
        }
        assert expected == set(COERCER_REGISTRY.keys())


# ------------------------------------------------------------------
# Connector integration with fixtures
# ------------------------------------------------------------------


def _asset_schema() -> SheetSchema:
    return SheetSchema(
        path=str(FIXTURES / "pmi_asset_registry_sample.xlsx"),
        sheet="Macchine",
        columns=[
            ColumnMapping(column="Codice", field="id", required=True),
            ColumnMapping(column="Nome", field="name", required=True),
            ColumnMapping(column="Tipo", field="type"),
            ColumnMapping(column="Ubicazione", field="location"),
            ColumnMapping(column="Costruttore", field="manufacturer"),
            ColumnMapping(column="Modello", field="model"),
            ColumnMapping(column="Matricola", field="serial_number"),
            ColumnMapping(column="Data Installazione", field="install_date", coerce="date_parse"),
            ColumnMapping(column="Criticità", field="criticality"),
            ColumnMapping(column="Padre", field="parent"),
        ],
    )


def _wo_schema(path: str) -> SheetSchema:
    return SheetSchema(
        path=path,
        sheet="OdL",
        columns=[
            ColumnMapping(column="ID", field="id", required=True),
            ColumnMapping(column="Tipo", field="type"),
            ColumnMapping(column="Priorità", field="priority"),
            ColumnMapping(column="Stato", field="status"),
            ColumnMapping(column="Codice Asset", field="asset_id", required=True),
            ColumnMapping(column="Descrizione", field="description"),
            ColumnMapping(column="Assegnato a", field="assigned_to"),
            ColumnMapping(
                column="Durata stimata (ore)", field="estimated_duration_hours", coerce="float_it"
            ),
            ColumnMapping(column="Data creazione", field="created_at", coerce="datetime_parse"),
        ],
        write_mode="append",
    )


@pytest.fixture()
def wo_file(tmp_path: Path) -> Path:
    """Copy blank workorders fixture to a temp dir for write tests."""
    src = FIXTURES / "workorders_blank.xlsx"
    dst = tmp_path / "workorders.xlsx"
    shutil.copy2(src, dst)
    return dst


class TestReadAssets:
    @pytest.mark.asyncio
    async def test_read_fixture(self) -> None:
        config = ExcelConnectorConfig(asset_registry=_asset_schema())
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        assets = await conn.read_assets()
        assert len(assets) >= 50
        first = assets[0]
        assert first.id == "P-001"
        assert first.name == "Pompa centrifuga linea 1"
        assert first.type == AssetType.ROTATING_EQUIPMENT
        assert first.criticality == Criticality.A
        assert first.install_date == date(2019, 3, 15)

    @pytest.mark.asyncio
    async def test_mixed_date_formats(self) -> None:
        config = ExcelConnectorConfig(asset_registry=_asset_schema())
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        assets = await conn.read_assets()
        by_id = {a.id: a for a in assets}
        # Italian dd/mm/yyyy
        assert by_id["P-001"].install_date == date(2019, 3, 15)
        # ISO yyyy-mm-dd
        assert by_id["P-002"].install_date == date(2020, 6, 22)

    @pytest.mark.asyncio
    async def test_excel_serial_date(self) -> None:
        config = ExcelConnectorConfig(asset_registry=_asset_schema())
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        assets = await conn.read_assets()
        by_id = {a.id: a for a in assets}
        assert by_id["PE-001"].install_date == date(2020, 1, 1)

    @pytest.mark.asyncio
    async def test_missing_optional_date(self) -> None:
        """Rows with empty date fields should parse with install_date=None."""
        config = ExcelConnectorConfig(asset_registry=_asset_schema())
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        assets = await conn.read_assets()
        by_id = {a.id: a for a in assets}
        assert by_id["TA-002"].install_date is None

    @pytest.mark.asyncio
    async def test_whitespace_stripped(self) -> None:
        config = ExcelConnectorConfig(asset_registry=_asset_schema())
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        assets = await conn.read_assets()
        by_id = {a.id: a for a in assets}
        # Asset has "  SP-001 " in the fixture — pydantic str_strip_whitespace handles it
        assert "SP-001" in by_id

    @pytest.mark.asyncio
    async def test_parent_field(self) -> None:
        config = ExcelConnectorConfig(asset_registry=_asset_schema())
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        assets = await conn.read_assets()
        by_id = {a.id: a for a in assets}
        assert by_id["C-002"].parent == "C-001"

    @pytest.mark.asyncio
    async def test_no_asset_registry_returns_empty(self) -> None:
        config = ExcelConnectorConfig(
            work_orders=_wo_schema(str(FIXTURES / "workorders_blank.xlsx")),
        )
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        assert await conn.read_assets() == []


class TestWriteWorkOrder:
    @pytest.mark.asyncio
    async def test_append_work_order(self, wo_file: Path) -> None:
        config = ExcelConnectorConfig(work_orders=_wo_schema(str(wo_file)))
        conn = ExcelCsvConnector(config=config)
        await conn.connect()

        wo = WorkOrder(
            id="WO-2026-001",
            type=WorkOrderType.CORRECTIVE,
            priority=Priority.HIGH,
            status=WorkOrderStatus.CREATED,
            asset_id="P-001",
            description="Sostituzione guarnizione pompa",
            assigned_to="Marco Rossi",
            estimated_duration_hours=2.5,
        )
        result = await conn.create_work_order(wo)
        assert result.id == "WO-2026-001"

        # Verify it was written to the file
        import openpyxl

        wb = openpyxl.load_workbook(str(wo_file))
        ws = wb["OdL"]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2  # header + 1 data row
        data_row = rows[1]
        assert data_row[0] == "WO-2026-001"
        assert data_row[4] == "P-001"
        wb.close()

    @pytest.mark.asyncio
    async def test_append_multiple(self, wo_file: Path) -> None:
        config = ExcelConnectorConfig(work_orders=_wo_schema(str(wo_file)))
        conn = ExcelCsvConnector(config=config)
        await conn.connect()

        for i in range(3):
            wo = WorkOrder(
                id=f"WO-2026-{i:03d}",
                type=WorkOrderType.CORRECTIVE,
                priority=Priority.MEDIUM,
                asset_id="P-001",
            )
            await conn.create_work_order(wo)

        wos = await conn.read_work_orders()
        assert len(wos) == 3

    @pytest.mark.asyncio
    async def test_no_write_mode_raises(self, tmp_path: Path) -> None:
        import openpyxl

        f = tmp_path / "readonly.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OdL"
        ws.append(["ID"])
        wb.save(str(f))
        wb.close()

        schema = SheetSchema(
            path=str(f),
            sheet="OdL",
            columns=[ColumnMapping(column="ID", field="id", required=True)],
            write_mode=None,
        )
        config = ExcelConnectorConfig(work_orders=schema)
        conn = ExcelCsvConnector(config=config)
        await conn.connect()

        wo = WorkOrder(id="WO-1", type=WorkOrderType.CORRECTIVE, asset_id="X")
        with pytest.raises(ConnectorConfigError, match="write_mode"):
            await conn.create_work_order(wo)


class TestErrorPaths:
    @pytest.mark.asyncio
    async def test_missing_file_raises(self) -> None:
        schema = SheetSchema(
            path="/nonexistent/assets.xlsx",
            sheet="Macchine",
            columns=[ColumnMapping(column="Codice", field="id", required=True)],
        )
        config = ExcelConnectorConfig(asset_registry=schema)
        conn = ExcelCsvConnector(config=config)
        with pytest.raises(ConnectorConfigError, match="not found"):
            await conn.connect()

    @pytest.mark.asyncio
    async def test_missing_required_column_raises(self, tmp_path: Path) -> None:
        """Schema expects a column that doesn't exist in the file."""
        import openpyxl

        f = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["WrongColumn"])
        ws.append(["value"])
        wb.save(str(f))
        wb.close()

        schema = SheetSchema(
            path=str(f),
            columns=[ColumnMapping(column="ExpectedColumn", field="id", required=True)],
        )
        config = ExcelConnectorConfig(asset_registry=schema)
        conn = ExcelCsvConnector(config=config)
        with pytest.raises(ConnectorSchemaError, match="Required columns missing"):
            await conn.connect()

    @pytest.mark.asyncio
    async def test_wrong_sheet_name_raises(self, tmp_path: Path) -> None:
        import openpyxl

        f = tmp_path / "sheets.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RealSheet"
        ws.append(["Col"])
        wb.save(str(f))
        wb.close()

        schema = SheetSchema(
            path=str(f),
            sheet="WrongSheet",
            columns=[ColumnMapping(column="Col", field="id")],
        )
        config = ExcelConnectorConfig(asset_registry=schema)
        conn = ExcelCsvConnector(config=config)
        with pytest.raises(ConnectorSchemaError, match="WrongSheet"):
            await conn.connect()


class TestHealthCheck:
    @pytest.mark.asyncio
    async def test_healthy(self) -> None:
        config = ExcelConnectorConfig(asset_registry=_asset_schema())
        conn = ExcelCsvConnector(config=config)
        health = await conn.health_check()
        assert health.status.value == "healthy"

    @pytest.mark.asyncio
    async def test_unhealthy_missing_file(self) -> None:
        schema = SheetSchema(
            path="/nonexistent/file.xlsx",
            columns=[ColumnMapping(column="X", field="id")],
        )
        config = ExcelConnectorConfig(asset_registry=schema)
        conn = ExcelCsvConnector(config=config)
        health = await conn.health_check()
        assert health.status.value == "unhealthy"


class TestCsvSupport:
    @pytest.mark.asyncio
    async def test_read_csv(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "assets.csv"
        csv_file.write_text(
            "Codice,Nome,Tipo\nP-001,Pompa 1,rotating_equipment\nP-002,Pompa 2,static_equipment\n",
            encoding="utf-8",
        )
        schema = SheetSchema(
            path=str(csv_file),
            columns=[
                ColumnMapping(column="Codice", field="id", required=True),
                ColumnMapping(column="Nome", field="name", required=True),
                ColumnMapping(column="Tipo", field="type"),
            ],
        )
        config = ExcelConnectorConfig(asset_registry=schema)
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        assets = await conn.read_assets()
        assert len(assets) == 2
        assert assets[0].id == "P-001"

    @pytest.mark.asyncio
    async def test_write_csv(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "odl.csv"
        schema = SheetSchema(
            path=str(csv_file),
            sheet="ignored",
            columns=[
                ColumnMapping(column="ID", field="id", required=True),
                ColumnMapping(column="Codice Asset", field="asset_id", required=True),
                ColumnMapping(column="Descrizione", field="description"),
            ],
            write_mode="append",
        )
        config = ExcelConnectorConfig(work_orders=schema)
        conn = ExcelCsvConnector(config=config)
        await conn.connect()

        wo = WorkOrder(
            id="WO-001", type=WorkOrderType.CORRECTIVE, asset_id="P-001", description="Test"
        )
        await conn.create_work_order(wo)

        content = csv_file.read_text(encoding="utf-8")
        assert "WO-001" in content
        assert "P-001" in content

    @pytest.mark.asyncio
    async def test_update_csv_persists_to_disk(self, tmp_path: Path) -> None:
        """A CSV-backed update must be written to disk, not just the cache —
        otherwise the change is lost on restart (the cache-only bug)."""
        csv_file = tmp_path / "odl.csv"
        schema = SheetSchema(
            path=str(csv_file),
            sheet="ignored",
            columns=[
                ColumnMapping(column="ID", field="id", required=True),
                ColumnMapping(column="Codice Asset", field="asset_id", required=True),
                ColumnMapping(column="Descrizione", field="description"),
            ],
            write_mode="append",
        )
        config = ExcelConnectorConfig(work_orders=schema)
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        await conn.create_work_order(
            WorkOrder(
                id="WO-001", type=WorkOrderType.CORRECTIVE, asset_id="P-001", description="old"
            )
        )
        await conn.update_work_order("WO-001", {"description": "new description"})

        # The on-disk file reflects the update (no duplicate row).
        content = csv_file.read_text(encoding="utf-8-sig")
        assert "new description" in content
        assert "old" not in content
        assert content.count("WO-001") == 1

        # A fresh connector reloads the updated value.
        conn2 = ExcelCsvConnector(config=config)
        await conn2.connect()
        reloaded = await conn2.read_work_orders()
        assert len(reloaded) == 1
        assert reloaded[0].description == "new description"

    @pytest.mark.asyncio
    async def test_csv_rewrite_atomic_on_failure(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """If the rewrite fails mid-way, the original file is left intact (the
        rewrite writes a temp sibling then atomically replaces the target)."""
        csv_file = tmp_path / "odl.csv"
        schema = SheetSchema(
            path=str(csv_file),
            sheet="ignored",
            columns=[
                ColumnMapping(column="ID", field="id", required=True),
                ColumnMapping(column="Codice Asset", field="asset_id", required=True),
                ColumnMapping(column="Descrizione", field="description"),
            ],
            write_mode="append",
        )
        config = ExcelConnectorConfig(work_orders=schema)
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        await conn.create_work_order(
            WorkOrder(
                id="WO-001", type=WorkOrderType.CORRECTIVE, asset_id="P-001", description="orig"
            )
        )
        before = csv_file.read_text(encoding="utf-8-sig")

        def _boom(wo: object, schema: object) -> dict[str, object]:
            raise RuntimeError("disk full mid-rewrite")

        monkeypatch.setattr(conn, "_work_order_to_row", _boom)
        with pytest.raises(RuntimeError):
            await conn.update_work_order("WO-001", {"description": "new"})

        # Original file unchanged — not truncated by the failed rewrite.
        assert csv_file.read_text(encoding="utf-8-sig") == before

    @pytest.mark.asyncio
    async def test_csv_formula_injection_neutralized_and_roundtrips(self, tmp_path: Path) -> None:
        """A field starting with a formula trigger is written with a leading
        apostrophe (so a spreadsheet treats it as text), and round-trips back
        to its original value through a fresh connector read."""
        csv_file = tmp_path / "odl.csv"
        schema = SheetSchema(
            path=str(csv_file),
            sheet="ignored",
            columns=[
                ColumnMapping(column="ID", field="id", required=True),
                ColumnMapping(column="Codice Asset", field="asset_id", required=True),
                ColumnMapping(column="Descrizione", field="description"),
            ],
            write_mode="append",
        )
        config = ExcelConnectorConfig(work_orders=schema)
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        await conn.create_work_order(
            WorkOrder(
                id="WO-001",
                type=WorkOrderType.CORRECTIVE,
                asset_id="P-001",
                description="=SUM(A1:A9)+cmd",
            )
        )
        # On disk the trigger is neutralized with a leading apostrophe.
        on_disk = csv_file.read_text(encoding="utf-8-sig")
        assert "'=SUM(A1:A9)+cmd" in on_disk
        assert ",=SUM" not in on_disk  # raw formula must not be present unguarded

        # A fresh connector reads back the original value (clean round-trip).
        conn2 = ExcelCsvConnector(config=config)
        await conn2.connect()
        reloaded = await conn2.read_work_orders()
        assert reloaded[0].description == "=SUM(A1:A9)+cmd"

    @pytest.mark.asyncio
    async def test_xlsx_formula_injection_neutralized_and_roundtrips(self, tmp_path: Path) -> None:
        """xlsx write neutralizes formula triggers (also stops openpyxl from
        storing the value as a real formula) and round-trips."""
        import openpyxl

        xlsx_file = tmp_path / "odl.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OdL"
        ws.append(["ID", "Codice Asset", "Descrizione"])
        wb.save(str(xlsx_file))
        wb.close()
        schema = SheetSchema(
            path=str(xlsx_file),
            sheet="OdL",
            columns=[
                ColumnMapping(column="ID", field="id", required=True),
                ColumnMapping(column="Codice Asset", field="asset_id", required=True),
                ColumnMapping(column="Descrizione", field="description"),
            ],
            write_mode="append",
        )
        config = ExcelConnectorConfig(work_orders=schema)
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        await conn.create_work_order(
            WorkOrder(
                id="WO-001",
                type=WorkOrderType.CORRECTIVE,
                asset_id="P-001",
                description="=HYPERLINK(0)",
            )
        )
        wb2 = openpyxl.load_workbook(str(xlsx_file))
        cell = wb2["OdL"].cell(row=2, column=3)
        wb2.close()
        # Stored as a text string (leading apostrophe), not a formula.
        assert cell.data_type != "f"
        assert cell.value == "'=HYPERLINK(0)"

        conn2 = ExcelCsvConnector(config=config)
        await conn2.connect()
        reloaded = await conn2.read_work_orders()
        assert reloaded[0].description == "=HYPERLINK(0)"


class TestRefresh:
    @pytest.mark.asyncio
    async def test_refresh_picks_up_new_data(self, tmp_path: Path) -> None:
        import openpyxl

        f = tmp_path / "assets.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Codice", "Nome"])
        ws.append(["P-001", "Pompa 1"])
        wb.save(str(f))
        wb.close()

        schema = SheetSchema(
            path=str(f),
            columns=[
                ColumnMapping(column="Codice", field="id", required=True),
                ColumnMapping(column="Nome", field="name", required=True),
            ],
        )
        config = ExcelConnectorConfig(asset_registry=schema)
        conn = ExcelCsvConnector(config=config)
        await conn.connect()
        assert len(await conn.read_assets()) == 1

        # Add a row externally
        wb = openpyxl.load_workbook(str(f))
        wb.active.append(["P-002", "Pompa 2"])
        wb.save(str(f))
        wb.close()

        conn.refresh()
        assert len(await conn.read_assets()) == 2


class TestFormulaGuardLossless:
    """_guard_formula / _strip_formula_guard must be a true inverse for every
    value class, including ones that already start with an apostrophe + trigger
    (the corruption case that the naive guard turned "'=approved" into
    "=approved" on read)."""

    @pytest.mark.parametrize(
        "value",
        [
            "=SUM(A1:A2)",
            "+1",
            "-1",
            "@cmd",
            "'=approved",  # literal apostrophe + trigger — must NOT be corrupted
            "''=x",  # double apostrophe + trigger
            "'note",  # apostrophe + non-trigger — left untouched
            "'",  # lone apostrophe
            "plain text",
            "",
        ],
    )
    def test_guard_strip_roundtrip(self, value: str) -> None:
        assert _strip_formula_guard(_guard_formula(value)) == value

    def test_trigger_is_neutralized_on_write(self) -> None:
        assert _guard_formula("=SUM(A1)") == "'=SUM(A1)"

    def test_ambiguous_literal_is_escaped_on_write(self) -> None:
        # The on-disk form gains an extra apostrophe so the strip is unambiguous.
        assert _guard_formula("'=approved") == "''=approved"
