"""ExcelCsvConnector — YAML-schema-driven Excel/CSV adapter.

Treats a directory of spreadsheets as a CMMS: reads assets and work
orders, appends new work orders.  Schema mapping is defined in YAML
so the user writes zero Python.
"""

from __future__ import annotations

import asyncio
import csv
import re
from datetime import UTC, date, datetime
from enum import StrEnum
from pathlib import Path
from typing import TYPE_CHECKING, Any, ClassVar

import structlog

from machina.connectors._entity_builders import dict_to_asset as _dict_to_asset
from machina.connectors._entity_builders import dict_to_failure_mode as _dict_to_failure_mode
from machina.connectors._entity_builders import dict_to_work_order as _dict_to_work_order
from machina.connectors.base import ConnectorHealth, ConnectorStatus, sandbox_aware
from machina.connectors.capabilities import Capability

if TYPE_CHECKING:
    from machina.connectors.docs.excel_schema import (
        ColumnMapping,
        ExcelConnectorConfig,
        SheetSchema,
    )
    from machina.domain.asset import Asset
    from machina.domain.failure_mode import FailureMode
    from machina.domain.work_order import WorkOrder
from machina.exceptions import (
    ConnectorConfigError,
    ConnectorError,
    ConnectorLockedError,
    ConnectorSchemaError,
)

logger = structlog.get_logger(__name__)


# ------------------------------------------------------------------
# Coercer registry — named functions referenced from YAML schemas
# ------------------------------------------------------------------


def _float_it(value: Any) -> float:
    """Parse a float, handling Italian decimal comma."""
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    return float(s)


def _int_it(value: Any) -> int:
    if isinstance(value, int):
        return value
    return int(_float_it(value))


_ITALIAN_DATE_RE = re.compile(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$")
_ISO_DATE_RE = re.compile(r"^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$")


def _date_parse(value: Any) -> date:
    """Parse a date from multiple formats: dd/mm/yyyy, yyyy-mm-dd, Excel serial."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return _excel_serial_to_date(value)
    s = str(value).strip()
    m = _ITALIAN_DATE_RE.match(s)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    m = _ISO_DATE_RE.match(s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # Last resort — try ISO parse
    return date.fromisoformat(s)


def _datetime_parse(value: Any) -> datetime:
    """Parse a datetime, delegating to _date_parse for date-only values."""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=UTC)
        return value
    if isinstance(value, (int, float)):
        d = _excel_serial_to_date(value)
        return datetime(d.year, d.month, d.day, tzinfo=UTC)
    s = str(value).strip()
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return dt
    except ValueError:
        d = _date_parse(s)
        return datetime(d.year, d.month, d.day, tzinfo=UTC)


def _excel_serial_to_date(serial: int | float) -> date:
    """Convert an Excel serial date number to a Python date."""
    # Excel epoch is 1899-12-30 (accounting for the Lotus 1-2-3 leap year bug)
    from datetime import timedelta

    base = date(1899, 12, 30)
    return base + timedelta(days=int(serial))


def _bool_it(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("1", "true", "yes", "sì", "si", "vero", "x"):
        return True
    if s in ("0", "false", "no", "falso", ""):
        return False
    msg = f"Cannot coerce {value!r} to bool"
    raise ValueError(msg)


def _strip(value: Any) -> str:
    return str(value).strip()


COERCER_REGISTRY: dict[str, Any] = {
    "float_it": _float_it,
    "int_it": _int_it,
    "date_parse": _date_parse,
    "italian_date": _date_parse,
    "datetime_parse": _datetime_parse,
    "bool_it": _bool_it,
    "strip": _strip,
}

_TYPE_COERCERS: dict[str, Any] = {
    "str": str,
    "int": _int_it,
    "float": _float_it,
    "date": _date_parse,
    "datetime": _datetime_parse,
    "bool": _bool_it,
}


# Leading characters a spreadsheet (Excel/LibreOffice) interprets as the start
# of a formula. A cell value beginning with one of these is a CSV/formula-
# injection vector when the exported file is opened in a spreadsheet app.
_FORMULA_PREFIXES: tuple[str, ...] = ("=", "+", "-", "@")


def _guard_formula(value: str) -> str:
    """Neutralize a leading formula trigger by prefixing an apostrophe.

    The apostrophe makes spreadsheets treat the cell as text (and stops
    openpyxl from storing it as a real formula). Reversed by
    :func:`_strip_formula_guard` on read so values round-trip unchanged.

    A value that *already* starts with an apostrophe followed by a trigger
    (or another apostrophe) is also escaped with an extra leading apostrophe,
    so the strip on read is unambiguous and the pair is a true inverse — a
    legitimate ``"'=approved"`` is not corrupted into ``"=approved"``.
    """
    head = value[:1]
    if head in _FORMULA_PREFIXES:
        return "'" + value
    if head == "'" and value[1:2] in (*_FORMULA_PREFIXES, "'"):
        return "'" + value
    return value


def _strip_formula_guard(value: str) -> str:
    """Reverse :func:`_guard_formula` so guarded values read back intact.

    Strips exactly one leading apostrophe when it is followed by a formula
    trigger or another apostrophe — the only shapes :func:`_guard_formula`
    ever produces — leaving genuine values like ``"'note"`` untouched.
    """
    if value[:1] == "'" and value[1:2] in (*_FORMULA_PREFIXES, "'"):
        return value[1:]
    return value


def _coerce_cell(value: Any, mapping: ColumnMapping) -> Any:
    """Coerce a single cell value according to its column mapping."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        if mapping.required:
            return None  # caller detects and flags the row
        return mapping.default
    if mapping.coerce and mapping.coerce in COERCER_REGISTRY:
        result = COERCER_REGISTRY[mapping.coerce](value)
    else:
        result = _TYPE_COERCERS.get(mapping.type, str)(value)
    if isinstance(result, str):
        result = _strip_formula_guard(result)
    return result


def _require_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as exc:
        raise ConnectorError(
            "openpyxl is required for Excel files. Install with: pip install machina-ai[excel]"
        ) from exc
    return openpyxl


# ------------------------------------------------------------------
# Row reading helpers
# ------------------------------------------------------------------


def _read_xlsx_rows(
    path: Path, sheet_name: str, schema: SheetSchema
) -> tuple[list[str], list[dict[str, Any]]]:
    """Read rows from an .xlsx file, returning (headers, list-of-row-dicts)."""
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except PermissionError as exc:
        raise ConnectorLockedError(f"File is locked by another process: {path.name}") from exc

    try:
        if sheet_name not in wb.sheetnames:
            raise ConnectorSchemaError(
                f"Sheet '{sheet_name}' not found in {path.name}. Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        data: list[dict[str, Any]] = []
        for row in rows_iter:
            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            data.append(row_dict)
        return headers, data
    finally:
        wb.close()


def _read_csv_rows(path: Path, schema: SheetSchema) -> tuple[list[str], list[dict[str, Any]]]:
    """Read rows from a CSV file."""
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        data = list(reader)
    return headers, data


def _validate_headers(headers: list[str], schema: SheetSchema, source: str) -> None:
    """Check that all required columns from the schema exist in the headers."""
    required_columns = {m.column for m in schema.columns if m.required}
    missing = required_columns - set(headers)
    if missing:
        raise ConnectorSchemaError(
            f"Required columns missing from {source}: {sorted(missing)}. "
            f"Available headers: {headers}"
        )


def _rows_to_dicts(
    raw_rows: list[dict[str, Any]],
    schema: SheetSchema,
    source: str,
) -> list[dict[str, Any]]:
    """Convert raw spreadsheet rows to coerced field dicts, skipping broken rows."""
    results: list[dict[str, Any]] = []
    for row_num, raw in enumerate(raw_rows, start=2):  # row 1 is header
        record: dict[str, Any] = {}
        broken = False
        for mapping in schema.columns:
            cell_value = raw.get(mapping.column)
            try:
                coerced = _coerce_cell(cell_value, mapping)
            except (ValueError, TypeError) as exc:
                if mapping.required:
                    logger.warning(
                        "broken_cell",
                        connector="ExcelCsvConnector",
                        source=source,
                        row_num=row_num,
                        column=mapping.column,
                        error_type=type(exc).__name__,
                        error=str(exc),
                    )
                    broken = True
                    break
                logger.debug(
                    "optional_cell_coercion_failed",
                    connector="ExcelCsvConnector",
                    source=source,
                    row_num=row_num,
                    column=mapping.column,
                    error=str(exc),
                )
                coerced = mapping.default
            if coerced is None and mapping.required:
                logger.warning(
                    "missing_required_field",
                    connector="ExcelCsvConnector",
                    source=source,
                    row_num=row_num,
                    column=mapping.column,
                    field=mapping.field,
                )
                broken = True
                break
            record[mapping.field] = coerced
        if not broken:
            results.append(record)
    return results


# ------------------------------------------------------------------
# Write helpers
# ------------------------------------------------------------------


def _append_xlsx_row(
    path: Path, sheet_name: str, schema: SheetSchema, row_data: dict[str, Any]
) -> None:
    """Append a single row to an .xlsx file."""
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(str(path))
    except PermissionError as exc:
        raise ConnectorLockedError(f"File is locked by another process: {path.name}") from exc
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append([m.column for m in schema.columns])

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
        wb[sheet_name].append([m.column for m in schema.columns])

    ws = wb[sheet_name]
    row_values = []
    for mapping in schema.columns:
        row_values.append(row_data.get(mapping.field))
    ws.append(row_values)
    wb.save(str(path))
    wb.close()


def _append_csv_row(path: Path, schema: SheetSchema, row_data: dict[str, Any]) -> None:
    """Append a single row to a CSV file."""
    file_exists = path.exists() and path.stat().st_size > 0
    columns = [m.column for m in schema.columns]
    with path.open("a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        if not file_exists:
            writer.writeheader()
        csv_row = {}
        for mapping in schema.columns:
            csv_row[mapping.column] = row_data.get(mapping.field, "")
        writer.writerow(csv_row)


# ------------------------------------------------------------------
# Connector
# ------------------------------------------------------------------


class ExcelCsvConnector:
    """Connector that treats Excel/CSV files as a CMMS substrate.

    Reads assets, work orders, and (optionally) a failure-mode catalog
    from spreadsheet files using a YAML schema mapping.  Writes new work
    orders by appending rows.  Multi-valued cells (asset failure-code
    linkage, failure-mode list fields) use a semicolon-delimited string,
    e.g. ``"BEAR-WEAR-01;SEAL-LEAK-01"``.

    Args:
        config: Parsed connector configuration.

    Example:
        ```python
        from machina.connectors.docs.excel import ExcelCsvConnector
        from machina.connectors.docs.excel_schema import ExcelConnectorConfig

        config = ExcelConnectorConfig.model_validate(yaml.safe_load(open("excel.yaml")))
        connector = ExcelCsvConnector(config=config)
        await connector.connect()
        assets = await connector.read_assets()
        ```
    """

    _BASE_CAPABILITIES: ClassVar[frozenset[Capability]] = frozenset(
        {
            Capability.READ_ASSETS,
            Capability.READ_WORK_ORDERS,
            Capability.CREATE_WORK_ORDER,
            Capability.UPDATE_WORK_ORDER,
        }
    )

    @property
    def capabilities(self) -> frozenset[Capability]:
        """Return capabilities based on configuration.

        Base capabilities are always available. ``READ_FAILURE_MODES`` is
        declared only when a ``failure_modes`` sheet is configured —
        unconfigured means not-declared, so capability discovery is a true
        signal of "has a failure-mode catalog".
        """
        return self._capabilities

    def __init__(self, *, config: ExcelConnectorConfig) -> None:
        self._config = config
        caps = set(self._BASE_CAPABILITIES)
        if config.failure_modes is not None:
            caps.add(Capability.READ_FAILURE_MODES)
        self._capabilities = frozenset(caps)
        self._connected = False
        self._asset_cache: list[Asset] = []
        self._wo_cache: list[WorkOrder] = []
        self._fm_cache: list[FailureMode] = []
        self._write_lock = asyncio.Lock()

    # ------------------------------------------------------------------
    # Lifecycle
    # ------------------------------------------------------------------

    async def connect(self) -> None:
        """Validate schemas against file headers and load initial data."""
        if self._config.asset_registry:
            self._validate_and_load_assets()
        if self._config.work_orders:
            self._validate_and_load_work_orders()
        if self._config.failure_modes:
            self._validate_and_load_failure_modes()
        self._connected = True
        logger.info(
            "connected",
            connector="ExcelCsvConnector",
            assets_loaded=len(self._asset_cache),
            work_orders_loaded=len(self._wo_cache),
            failure_modes_loaded=len(self._fm_cache),
        )

    async def disconnect(self) -> None:
        """Release caches."""
        self._asset_cache.clear()
        self._wo_cache.clear()
        self._fm_cache.clear()
        self._connected = False

    async def health_check(self) -> ConnectorHealth:
        """Check that configured files are accessible."""
        issues: list[str] = []
        for label, schema in [
            ("asset_registry", self._config.asset_registry),
            ("work_orders", self._config.work_orders),
            ("failure_modes", self._config.failure_modes),
        ]:
            if schema is None:
                continue
            p = Path(schema.path)
            if not p.exists():
                issues.append(f"{label}: file not found ({p})")
        if issues:
            return ConnectorHealth(
                status=ConnectorStatus.UNHEALTHY,
                message="; ".join(issues),
            )
        return ConnectorHealth(status=ConnectorStatus.HEALTHY, message="All files accessible")

    # ------------------------------------------------------------------
    # Read operations
    # ------------------------------------------------------------------

    async def read_assets(self) -> list[Asset]:
        """Return assets from the asset registry spreadsheet."""
        if self._config.asset_registry is None:
            return []
        return list(self._asset_cache)

    async def read_work_orders(self) -> list[WorkOrder]:
        """Return work orders from the work-order spreadsheet."""
        if self._config.work_orders is None:
            return []
        return list(self._wo_cache)

    async def read_failure_modes(self) -> list[FailureMode]:
        """Return failure modes from the failure-modes spreadsheet.

        Returns an empty list when no ``failure_modes`` sheet is
        configured (the capability is then not declared either).

        Raises:
            ConnectorError: If a sheet is configured but the connector is
                not connected — matching the cross-substrate harvest
                contract, so a configured catalog never silently reads
                as "no failure-mode data configured".
        """
        if self._config.failure_modes is None:
            return []
        if not self._connected:
            raise ConnectorError("Not connected — call connect() before reading")
        return list(self._fm_cache)

    # ------------------------------------------------------------------
    # Write operations
    # ------------------------------------------------------------------

    @sandbox_aware
    async def create_work_order(self, work_order: WorkOrder) -> WorkOrder:
        """Append a new work order row to the spreadsheet."""
        schema = self._config.work_orders
        if schema is None:
            raise ConnectorConfigError("No work_orders schema configured for writing")
        if schema.write_mode is None:
            raise ConnectorConfigError("work_orders schema has no write_mode configured")

        row_data = self._work_order_to_row(work_order, schema)
        path = Path(schema.path)

        async with self._write_lock:
            await asyncio.to_thread(self._write_row, path, schema, row_data)
        self._wo_cache.append(work_order)
        logger.info(
            "work_order_created",
            connector="ExcelCsvConnector",
            work_order_id=work_order.id,
            asset_id=work_order.asset_id,
        )
        return work_order

    @sandbox_aware
    async def update_work_order(self, work_order_id: str, updates: dict[str, Any]) -> WorkOrder:
        """Update a work order in cache and persist to file.

        When ``write_mode`` is configured, a full rewrite from cache is
        performed for both xlsx and csv files, so the change is durable
        across restarts. When no ``write_mode`` is set, the update is kept
        in cache only.
        """
        for wo in self._wo_cache:
            if wo.id == work_order_id:
                schema = self._config.work_orders
                # Serialise the cache mutation together with the rewrite under
                # the write lock: the rewrite reads the whole cache from a worker
                # thread, so a concurrent update mutating the cache must not
                # interleave with it.
                async with self._write_lock:
                    for key, value in updates.items():
                        if hasattr(wo, key):
                            setattr(wo, key, value)
                    if schema and schema.write_mode:
                        # Full rewrite from cache for both xlsx and csv so the
                        # change is durable, not cache-only (lost on restart).
                        await asyncio.to_thread(self._rewrite_work_orders, schema)
                if not (schema and schema.write_mode):
                    logger.warning(
                        "update_not_persisted",
                        connector="ExcelCsvConnector",
                        work_order_id=work_order_id,
                        hint="no write_mode configured — update kept in cache only",
                    )
                logger.info(
                    "work_order_updated",
                    connector="ExcelCsvConnector",
                    work_order_id=work_order_id,
                )
                return wo
        raise ConnectorError(f"Work order '{work_order_id}' not found in cache")

    def _rewrite_work_orders(self, schema: SheetSchema) -> None:
        """Rewrite all work orders from cache, dispatching by file format."""
        path = Path(schema.path)
        if path.suffix.lower() == ".csv":
            self._rewrite_csv(path, schema)
        else:
            self._rewrite_xlsx(path, schema)

    def _rewrite_xlsx(self, path: Path, schema: SheetSchema) -> None:
        """Rewrite all work orders to the xlsx file from cache.

        Writes to a temp sibling then atomically replaces the target, so a
        crash or error mid-write cannot truncate the existing file.
        """
        openpyxl = _require_openpyxl()
        tmp = path.with_name(path.name + ".tmp")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = schema.sheet
            ws.append([m.column for m in schema.columns])
            for wo in self._wo_cache:
                row_data = self._work_order_to_row(wo, schema)
                row_values = [row_data.get(m.field) for m in schema.columns]
                ws.append(row_values)
            wb.save(str(tmp))
            wb.close()
            tmp.replace(path)
        except Exception:
            # Don't leave a partial/orphaned temp file behind on failure.
            tmp.unlink(missing_ok=True)
            raise

    def _rewrite_csv(self, path: Path, schema: SheetSchema) -> None:
        """Rewrite all work orders to the CSV file from cache (header + rows).

        Writes to a temp sibling then atomically replaces the target, so a
        crash or error mid-write cannot truncate the existing file.
        """
        columns = [m.column for m in schema.columns]
        tmp = path.with_name(path.name + ".tmp")
        try:
            with tmp.open("w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                for wo in self._wo_cache:
                    row_data = self._work_order_to_row(wo, schema)
                    writer.writerow({m.column: row_data.get(m.field, "") for m in schema.columns})
            tmp.replace(path)
        except Exception:
            # Don't leave a partial/orphaned temp file behind on failure.
            tmp.unlink(missing_ok=True)
            raise

    # ------------------------------------------------------------------
    # Cache refresh (called by watcher)
    # ------------------------------------------------------------------

    def refresh(self) -> None:
        """Re-read files and update caches. Called by the watcher on file changes.

        All-or-nothing: if any sheet fails to load mid-refresh (file
        mid-save, locked, header change), every cache is restored to its
        pre-refresh snapshot so assets and the failure-mode catalog never
        end up mutually inconsistent.
        """
        snapshot = (
            list(self._asset_cache),
            list(self._wo_cache),
            list(self._fm_cache),
        )
        try:
            if self._config.asset_registry:
                self._validate_and_load_assets()
            if self._config.work_orders:
                self._validate_and_load_work_orders()
            if self._config.failure_modes:
                self._validate_and_load_failure_modes()
        except Exception:
            self._asset_cache, self._wo_cache, self._fm_cache = snapshot
            raise
        logger.info(
            "cache_refreshed",
            connector="ExcelCsvConnector",
            assets=len(self._asset_cache),
            work_orders=len(self._wo_cache),
            failure_modes=len(self._fm_cache),
        )

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _load_sheet_dicts(self, schema: SheetSchema, label: str) -> list[dict[str, Any]]:
        """Shared exists-check → read → validate → parse pipeline for one sheet."""
        for col in schema.columns:
            if col.coerce and col.coerce not in COERCER_REGISTRY:
                raise ConnectorConfigError(
                    f"Unknown coerce '{col.coerce}' for column '{col.column}' "
                    f"({label}) — known coercers: {sorted(COERCER_REGISTRY)}. "
                    "For plain type conversion use the 'type' field instead."
                )
        path = Path(schema.path)
        if not path.exists():
            raise ConnectorConfigError(f"{label} file not found: {path}")
        headers, raw_rows = self._read_file(path, schema)
        _validate_headers(headers, schema, str(path))
        return _rows_to_dicts(raw_rows, schema, str(path))

    def _validate_and_load_assets(self) -> None:
        schema = self._config.asset_registry
        assert schema is not None
        dicts = self._load_sheet_dicts(schema, "Asset registry")
        self._asset_cache = [_dict_to_asset(d) for d in dicts]

    def _validate_and_load_failure_modes(self) -> None:
        schema = self._config.failure_modes
        assert schema is not None
        dicts = self._load_sheet_dicts(schema, "Failure modes")
        self._fm_cache = [_dict_to_failure_mode(d) for d in dicts]

    def _validate_and_load_work_orders(self) -> None:
        schema = self._config.work_orders
        assert schema is not None
        if not Path(schema.path).exists() and schema.write_mode is not None:
            self._wo_cache = []
            return
        dicts = self._load_sheet_dicts(schema, "Work order")
        self._wo_cache = [_dict_to_work_order(d) for d in dicts]

    @staticmethod
    def _read_file(path: Path, schema: SheetSchema) -> tuple[list[str], list[dict[str, Any]]]:
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            return _read_xlsx_rows(path, schema.sheet, schema)
        if suffix == ".csv":
            return _read_csv_rows(path, schema)
        raise ConnectorConfigError(f"Unsupported file format: {suffix}")

    @staticmethod
    def _write_row(path: Path, schema: SheetSchema, row_data: dict[str, Any]) -> None:
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            _append_xlsx_row(path, schema.sheet, schema, row_data)
        elif suffix == ".csv":
            _append_csv_row(path, schema, row_data)
        else:
            raise ConnectorConfigError(f"Unsupported file format for writing: {suffix}")

    @staticmethod
    def _work_order_to_row(wo: WorkOrder, schema: SheetSchema) -> dict[str, Any]:
        wo_dict = wo.model_dump()
        row: dict[str, Any] = {}
        for mapping in schema.columns:
            value = wo_dict.get(mapping.field)
            if isinstance(value, (datetime, date)):
                value = value.isoformat()
            elif isinstance(value, StrEnum):
                value = value.value
            if isinstance(value, str):
                value = _guard_formula(value)
            row[mapping.field] = value
        return row
